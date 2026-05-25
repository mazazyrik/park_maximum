import io
from datetime import datetime
from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Order


def export_xlsx(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Заявки'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='282828')

    headers = [
        '№', 'ФИО', 'Телефон', 'Откуда', 'Куда', 'Тариф', 'Автомобиль',
        'Дата поездки', 'Отч. документы', 'Расстояние (км)',
        'Стоимость (руб)', 'Статус', 'Дата заявки',
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for obj in queryset.select_related('tariff', 'car'):
        ws.append([
            obj.pk,
            obj.fio,
            obj.phone,
            obj.from_address,
            obj.to_address,
            obj.tariff.name if obj.tariff else '',
            obj.car.name if obj.car else '',
            obj.trip_datetime.strftime('%d.%m.%Y %H:%M') if obj.trip_datetime else '',
            'Да' if obj.need_docs else 'Нет',
            obj.distance_km,
            float(obj.estimated_cost),
            obj.get_status_display(),
            obj.created_at.strftime('%d.%m.%Y %H:%M'),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'orders_{datetime.now().strftime("%Y-%m")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


export_xlsx.short_description = 'Выгрузить в XLSX'


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = (
        'pk', 'fio', 'phone', 'from_address_short', 'to_address_short',
        'tariff', 'trip_datetime', 'need_docs', 'distance_km',
        'estimated_cost', 'status_badge', 'created_at',
    )
    list_filter = (
        'status',
        'tariff',
        'need_docs',
        ('trip_datetime', admin.DateFieldListFilter),
        ('created_at', admin.DateFieldListFilter),
    )
    search_fields = ('fio', 'phone', 'from_address', 'to_address')
    date_hierarchy = 'created_at'
    readonly_fields = ('created_at', 'updated_at', 'estimated_cost', 'distance_km')
    actions = [export_xlsx]
    list_per_page = 50

    fieldsets = (
        ('Маршрут', {
            'fields': ('from_address', 'to_address', 'distance_km', 'estimated_cost'),
        }),
        ('Тариф', {
            'fields': ('tariff', 'car', 'trip_datetime', 'need_docs'),
        }),
        ('Клиент', {
            'fields': ('fio', 'phone'),
        }),
        ('Статус', {
            'fields': ('status', 'notes'),
        }),
        ('Служебное', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',),
        }),
    )

    @admin.display(description='Откуда')
    def from_address_short(self, obj):
        return obj.from_address[:40] + '…' if len(obj.from_address) > 40 else obj.from_address

    @admin.display(description='Куда')
    def to_address_short(self, obj):
        return obj.to_address[:40] + '…' if len(obj.to_address) > 40 else obj.to_address

    @admin.display(description='Статус')
    def status_badge(self, obj):
        colors = {
            'new': '#f5a623',
            'accepted': '#4a90e2',
            'in_progress': '#7ed321',
            'completed': '#417505',
            'cancelled': '#d0021b',
        }
        color = colors.get(obj.status, '#999')
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:10px;font-size:11px">{}</span>',
            color,
            obj.get_status_display(),
        )
